"""
读取Excel文件

Version: 0.1
Author: 骆昊
Date: 2018-03-26
"""

from openpyxl import load_workbook
from openpyxl import Workbook

workbook = load_workbook('C:/Users/chenchen/Documents/GitHub/Python-100-Days-1/Day01-15/code/Day15/res/学生明细表.xlsx')
# 打印所有的工作表的名称
print(workbook.sheetnames)
sheet = workbook[workbook.sheetnames[0]]
print(sheet.title)
for row in range(2, 7):
    # 从列 'A'（ASCII 65）到列 'E'（ASCII 69）遍历：
    for col in range(65, 70):
        # 这部分代码使用 chr 函数将整数 col 转换为对应的 ASCII 字符
        # 例如，如果 col 是 65，row 是 2，那么cell_index 将成为 'A2'
        cell_index = chr(col) + str(row)
        print(sheet[cell_index].value, end='/t')
    print()
